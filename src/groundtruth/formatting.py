"""Formatting configuration for Groundtruth XLSX output."""

from pathlib import Path
from typing import Any

import yaml
from openpyxl.styles import Font, PatternFill

# path to default formatting config
DEFAULTS_PATH = Path(__file__).parent / "defaults" / "formatting.yaml"


def load_formatting_config(user_config_path: Path | None = None) -> dict[str, Any]:
    """
    Load formatting configuration with defaults and optional user overrides.

    Args:
        user_config_path: Optional path to user's formatting config file

    Returns:
        Merged configuration dictionary
    """
    # load defaults
    with open(DEFAULTS_PATH, encoding="utf-8") as f:
        config = yaml.safe_load(f) or {}

    # overlay user config if provided
    if user_config_path and user_config_path.exists():
        with open(user_config_path, encoding="utf-8") as f:
            user_config = yaml.safe_load(f) or {}
        config = _deep_merge(config, user_config)

    return config


def _deep_merge(base: dict, overlay: dict) -> dict:
    """Deep merge overlay into base, returning new dict."""
    result = base.copy()
    for key, value in overlay.items():
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            result[key] = _deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def get_significance_styles(config: dict[str, Any]) -> tuple[dict[str, PatternFill], dict[str, Font]]:
    """
    Get significance fill and font styles from config.

    Returns:
        Tuple of (fill_dict, font_dict) keyed by significance level string
    """
    fills = {}
    fonts = {}

    sig_config = config.get("colors", {}).get("significance", {})
    for level in ["1", "2", "3", "4", "5"]:
        level_config = sig_config.get(level, {})
        bg = level_config.get("background", "FFFFFF")
        text = level_config.get("text", "000000")
        bold = level_config.get("bold", False)

        fills[level] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        fonts[level] = Font(bold=bold, color=text)

    return fills, fonts


def get_status_styles(config: dict[str, Any]) -> dict[str, PatternFill]:
    """Get status fill styles from config."""
    fills = {}
    status_config = config.get("colors", {}).get("status", {})

    # map config keys to display values
    key_map = {
        "agreed": "Agreed",
        "needs-clarification": "Needs Clarification",
        "unresolved": "Unresolved",
    }

    for config_key, display_key in key_map.items():
        level_config = status_config.get(config_key, {})
        bg = level_config.get("background", "FFFFFF")
        fills[display_key] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    return fills


def get_agreement_styles(config: dict[str, Any]) -> dict[str, PatternFill]:
    """Get agreement fill styles from config."""
    fills = {}
    agreement_config = config.get("colors", {}).get("agreement", {})

    # map config keys to display values
    key_map = {
        "yes": "Yes",
        "partial": "Partial",
        "no": "No",
    }

    for config_key, display_key in key_map.items():
        level_config = agreement_config.get(config_key, {})
        bg = level_config.get("background", "FFFFFF")
        fills[display_key] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    return fills


def get_header_styles(config: dict[str, Any]) -> tuple[PatternFill, Font]:
    """Get header fill and font styles from config."""
    header_config = config.get("colors", {}).get("header", {})
    bg = header_config.get("background", "4472C4")
    text = header_config.get("text", "FFFFFF")

    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(bold=True, color=text)

    return fill, font


def get_column_widths(config: dict[str, Any]) -> dict[str, int]:
    """Get column width settings from config."""
    return config.get("column_widths", {})
