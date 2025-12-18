"""XLSX generator for Groundtruth reports."""

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from groundtruth.formatting import (
    get_agreement_styles,
    get_header_styles,
    get_significance_styles,
    get_status_styles,
    load_formatting_config,
)
from groundtruth.models import DEFAULT_PARTICIPANTS

# Standard styles
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(wrap_text=True, vertical="top", horizontal="center")
HEADER_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column order:
# Category, Significance, Status, Title, Description, Decision, [Agreed], Notes, Meeting Date, Ref
COLUMN_ORDER = [
    ("Category", 20, False),
    ("Significance", 10, True),
    ("Status", 18, True),
    ("Title", 28, False),
    ("Description", 55, False),
    ("Decision", 30, False),
    # Agreed columns inserted dynamically
    # Notes, Meeting Date, Meeting Reference appended after
]

TRAILING_COLUMNS = [
    ("Notes", 45, False),
    ("Meeting Date", 12, True),
    ("Meeting Reference", 28, False),
]


def get_column_config(participants: list[str]) -> dict[str, dict[str, Any]]:
    """Get column configuration based on participants."""
    config: dict[str, dict[str, Any]] = {}
    col_letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    col_idx = 0

    # base columns before agreed columns
    for name, width, center in COLUMN_ORDER:
        config[col_letters[col_idx]] = {"name": name, "width": width, "center": center}
        col_idx += 1

    # agreed columns for each participant
    for participant in participants:
        config[col_letters[col_idx]] = {
            "name": f"{participant} Agreed",
            "width": 12,
            "center": True,
        }
        col_idx += 1

    # trailing columns
    for name, width, center in TRAILING_COLUMNS:
        config[col_letters[col_idx]] = {"name": name, "width": width, "center": center}
        col_idx += 1

    return config


def apply_cell_style(
    ws: Worksheet,
    row_idx: int,
    col_idx: int,
    value: str,
    is_header: bool,
    col_config: dict[str, Any],
    participants: list[str],
    styles: dict[str, Any],
) -> None:
    """Apply styling to a cell."""
    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    cell.border = THIN_BORDER

    if is_header:
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = HEADER_ALIGNMENT
        return

    # data row styling
    cell.alignment = CENTER_ALIGNMENT if col_config.get("center") else WRAP_ALIGNMENT

    # significance column (B = 2)
    if col_idx == 2 and value in styles["significance_fills"]:
        cell.fill = styles["significance_fills"][value]
        cell.font = styles["significance_fonts"].get(value, Font())

    # status column (C = 3)
    if col_idx == 3 and value in styles["status_fills"]:
        cell.fill = styles["status_fills"][value]

    # agreement columns (start at G = 7)
    agreement_start = 7
    agreement_end = agreement_start + len(participants)
    if agreement_start <= col_idx < agreement_end and value in styles["agreement_fills"]:
        cell.fill = styles["agreement_fills"][value]
        if value in ["No", "Partial"]:
            cell.font = Font(bold=True)


def add_summary_sheet(
    wb: Workbook,
    rows: list[list[str]],
    participants: list[str],
    file_metadata: dict[str, dict[str, Any]] | None = None,
) -> None:
    """Add summary sheet with input files and decision counts.

    Args:
        wb: Workbook to add sheet to
        rows: Decision data rows including header
        participants: List of participant names (global)
        file_metadata: Optional dict mapping filename to metadata:
            {filename: {"participants": [...], "deciders": [...]}}
    """
    ws = wb.create_sheet(title="Summary")

    # title
    ws["A1"] = "Decisions Summary"
    ws["A1"].font = Font(bold=True, size=16)

    data_rows = rows[1:]  # skip header

    # calculate column indices for meeting reference
    ref_col_idx = 6 + len(participants) + 1  # Notes=6+n, Date=7+n, Ref=8+n

    # group decisions by input file
    file_decisions: dict[str, int] = {}
    for row in data_rows:
        if len(row) > ref_col_idx and row[ref_col_idx]:
            filename = row[ref_col_idx]
            file_decisions[filename] = file_decisions.get(filename, 0) + 1

    # input files section with header row
    ws["A3"] = "Input File"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "Decisions"
    ws["B3"].font = Font(bold=True)
    ws["C3"] = "Participants"
    ws["C3"].font = Font(bold=True)
    ws["D3"] = "Deciders"
    ws["D3"].font = Font(bold=True)

    row_num = 4
    for filename in sorted(file_decisions.keys()):
        ws[f"A{row_num}"] = filename
        ws[f"B{row_num}"] = file_decisions[filename]

        # add per-file metadata if available
        if file_metadata and filename in file_metadata:
            meta = file_metadata[filename]
            if "participants" in meta:
                ws[f"C{row_num}"] = ", ".join(meta["participants"])
            if "deciders" in meta:
                ws[f"D{row_num}"] = ", ".join(meta["deciders"])
        row_num += 1

    row_num += 1

    # total decision count
    ws[f"A{row_num}"] = "Total Decisions"
    ws[f"A{row_num}"].font = Font(bold=True)
    ws[f"B{row_num}"] = len(data_rows)

    # set column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


def _parse_markdown_line(line: str) -> tuple[str, Font]:
    """Parse a markdown line and return (text, font) for Excel."""
    line = line.strip()

    # H1: # Header
    if line.startswith("# "):
        return line[2:], Font(bold=True, size=14)

    # H2: ## Header
    if line.startswith("## "):
        return line[3:], Font(bold=True, size=12)

    # H3: ### Header
    if line.startswith("### "):
        return line[4:], Font(bold=True, size=11)

    # Bold: **text** or __text__
    if line.startswith("**") and line.endswith("**"):
        return line[2:-2], Font(bold=True)
    if line.startswith("__") and line.endswith("__"):
        return line[2:-2], Font(bold=True)

    # Italic: *text* or _text_
    if line.startswith("*") and line.endswith("*") and not line.startswith("**"):
        return line[1:-1], Font(italic=True)
    if line.startswith("_") and line.endswith("_") and not line.startswith("__"):
        return line[1:-1], Font(italic=True)

    # Bullet point: - item or * item
    if line.startswith("- ") or line.startswith("* "):
        return "  " + line, Font()

    return line, Font()


def add_attribution_sheet(
    wb: Workbook,
    user_name: str | None = None,
    user_email: str | None = None,
    decision_framework: str | None = None,
) -> None:
    """Add attribution sheet to workbook."""
    ws = wb.create_sheet(title="Produced By")

    # set wider columns for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60

    # title - merged across columns
    ws.merge_cells("A1:C1")
    ws["A1"] = "Groundtruth"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="left")

    # metadata section
    ws["A3"] = "Tool"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "groundtruth"

    ws["A4"] = "Repository"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "https://github.com/rsnodgrass/groundtruth"

    ws["A5"] = "Generated"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = 6
    if user_name:
        ws[f"A{row}"] = "Author"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = user_name
        row += 1

    if user_email:
        ws[f"A{row}"] = "Email"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = user_email
        row += 1

    # quick reference section - merged header
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Quick Reference: Significance & Agreement Standards"
    ws[f"A{row}"].font = Font(bold=True, size=14)

    row += 2
    significance_ref = [
        ("Level", "Label", "Agreement Standard"),
        ("1", "Critical", "ALL must explicitly agree; any ambiguity = No"),
        ("2", "Extremely Important", "ALL must explicitly agree; any ambiguity = No"),
        ("3", "Important", "Any hint of misalignment = Partial or No"),
        ("4", "Moderate", "Clear alignment; slight ambiguity OK"),
        ("5", "Same Page", "General alignment sufficient"),
    ]

    for i, (level, label, standard) in enumerate(significance_ref):
        ws[f"A{row}"] = level
        ws[f"B{row}"] = label
        ws[f"C{row}"] = standard
        if i == 0:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"C{row}"].font = Font(bold=True)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Agreement: Yes (explicit), Partial (hedged), No (silent/parked/disagreed)"
    ws[f"A{row}"].font = Font(italic=True)

    # user's custom framework
    if decision_framework:
        row += 3
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = "Decision Framework"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = "User-provided framework that guided extraction:"
        ws[f"A{row}"].font = Font(italic=True, color="666666")
        row += 2

        # parse markdown and write framework content
        framework_lines = decision_framework.strip().split("\n")
        for line in framework_lines:
            if not line.strip():
                row += 1
                continue

            text, font = _parse_markdown_line(line)
            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"] = text
            ws[f"A{row}"].font = font
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            row += 1


def generate_xlsx(
    rows: list[list[str]],
    output_path: Path,
    participants: list[str] | None = None,
    user_name: str | None = None,
    user_email: str | None = None,
    decision_framework: str | None = None,
    format_config_path: Path | None = None,
    file_metadata: dict[str, dict[str, Any]] | None = None,
) -> int:
    """
    Generate formatted XLSX from decision data.

    Args:
        rows: List of rows including header
        output_path: Path to save XLSX
        participants: List of participant names
        user_name: Optional user name for attribution
        user_email: Optional user email for attribution
        decision_framework: Optional custom prompt that guided decision extraction
        format_config_path: Optional path to formatting config YAML
        file_metadata: Optional dict mapping filename to per-file metadata:
            {filename: {"participants": [...], "deciders": [...]}}

    Returns:
        Number of decisions (excluding header)
    """
    if participants is None:
        participants = DEFAULT_PARTICIPANTS

    # load formatting config (defaults + optional user overrides)
    format_config = load_formatting_config(format_config_path)

    # build styles from config
    significance_fills, significance_fonts = get_significance_styles(format_config)
    status_fills = get_status_styles(format_config)
    agreement_fills = get_agreement_styles(format_config)
    header_fill, header_font = get_header_styles(format_config)

    styles = {
        "significance_fills": significance_fills,
        "significance_fonts": significance_fonts,
        "status_fills": status_fills,
        "agreement_fills": agreement_fills,
        "header_fill": header_fill,
        "header_font": header_font,
    }

    # try to get user info from environment
    if user_name is None:
        user_name = os.environ.get("USER") or os.environ.get("USERNAME")
    if user_email is None:
        user_email = os.environ.get("EMAIL")

    # sort data by Category, then Significance
    header = rows[0]
    data = rows[1:]

    # find significance column index (should be 2 in new order)
    sig_col_idx = 2
    data.sort(
        key=lambda x: (
            x[0] if x else "",
            int(x[sig_col_idx]) if len(x) > sig_col_idx and x[sig_col_idx].isdigit() else 99,
        )
    )
    sorted_rows = [header] + data

    # create workbook
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")
    ws.title = "Decisions"

    # get column configuration
    col_config = get_column_config(participants)

    # set column widths
    for col_letter, config in col_config.items():
        ws.column_dimensions[col_letter].width = config["width"]

    # write data with styling
    for row_idx, row_data in enumerate(sorted_rows, 1):
        is_header = row_idx == 1
        for col_idx, value in enumerate(row_data, 1):
            col_letter = chr(ord("A") + col_idx - 1)
            config = col_config.get(col_letter, {"center": False})
            apply_cell_style(ws, row_idx, col_idx, value, is_header, config, participants, styles)

    # freeze header and set row heights
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35
    for row_idx in range(2, len(sorted_rows) + 1):
        ws.row_dimensions[row_idx].height = 60

    # add summary sheet
    add_summary_sheet(wb, sorted_rows, participants, file_metadata)

    # add attribution sheet
    add_attribution_sheet(wb, user_name, user_email, decision_framework)

    # save
    wb.save(output_path)
    return len(sorted_rows) - 1


def generate_from_csv(
    csv_path: Path,
    output_path: Path | None = None,
    participants: list[str] | None = None,
    decision_framework: str | None = None,
) -> tuple[Path, int]:
    """
    Generate XLSX from a CSV file.

    Args:
        csv_path: Path to CSV file
        output_path: Optional output path (defaults to same name with .xlsx)
        participants: List of participant names
        decision_framework: Optional custom prompt that guided decision extraction

    Returns:
        Tuple of (output path, decision count)
    """
    if output_path is None:
        output_path = csv_path.with_suffix(".xlsx")

    # read CSV
    rows: list[list[str]] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)

    count = generate_xlsx(rows, output_path, participants, decision_framework=decision_framework)
    return output_path, count


def get_csv_header(participants: list[str] | None = None) -> list[str]:
    """Get the standard CSV header row."""
    if participants is None:
        participants = DEFAULT_PARTICIPANTS

    return [
        "Category",
        "Significance",
        "Status",
        "Title",
        "Description",
        "Decision",
        *[f"{p} Agreed" for p in participants],
        "Notes",
        "Meeting Date",
        "Meeting Reference",
    ]
